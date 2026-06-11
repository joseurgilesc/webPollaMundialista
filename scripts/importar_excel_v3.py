#!/usr/bin/env python3
"""
Parser v3 para pollas formato SCC-variante (Emiliano, Mónica).

Diferencias con SCC:
- Grupos: columnas B/E en vez de A/D (filas 5-40)
- Bracket izq: I(slots)/J(16avos)/L(8avos)/N(cuartos)/P(semis)
- Bracket der: AE(slots)/AD(16avos)/AB(8avos)/Z(cuartos)/X(semis)
- Participante: G2
- Finales: CAMPEÓN en S16/S17

Uso: python3 importar_excel_v3.py <archivo.xlsx> [...]
"""

import json, os, re, sys
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl no instalado.")
    sys.exit(1)

def limpiar(texto):
    return " ".join(str(texto).split()).strip() if texto else ""

def normalizar_equipo(texto):
    return limpiar(texto)

def es_slot_code(texto):
    t = limpiar(texto).upper()
    return bool(re.match(r'^[123]\s*[A-L]+$', t))

def parsear_excel_v3(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    grilla = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 60):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    grilla[(cell.row, cell.column_letter)] = val
    
    # ── Participante ──
    participante = grilla.get((2, "G"), "Sin nombre")
    if participante == "Sin nombre":
        participante = grilla.get((2, "F"), "Sin nombre")
    
    # ── Grupos: search B+E columns for "GRUPO" ──
    grupos = {}
    for col_grupo, col_pos in [("B", "C"), ("E", "F")]:
        for fila in range(1, 50):
            val = grilla.get((fila, col_grupo), "")
            if val.upper().startswith("GRUPO"):
                letra = val.replace("GRUPO", "").strip()
                if not letra or len(letra) > 2:
                    continue
                equipos = {}
                for offset in range(1, 5):
                    f = fila + offset
                    eq = limpiar(grilla.get((f, col_grupo), ""))
                    pos_str = grilla.get((f, col_pos), "")
                    if not eq:
                        continue
                    try:
                        pos = int(float(pos_str))
                    except (ValueError, TypeError):
                        pos = None
                    equipos[eq] = pos
                if equipos:
                    grupos[letra] = equipos
    
    # ── Bracket LEFT: I(slots)/J(16avos)/L(8avos)/N(cuartos)/P(semis) ──
    bracket_izq = {"16avos": [], "8avos": [], "cuartos": [], "semis": []}
    for fila in range(7, 40, 2):
        slot = grilla.get((fila, "I"), "")
        eq = grilla.get((fila, "J"), "")
        if slot and es_slot_code(slot):
            bracket_izq["16avos"].append({"slot": slot, "equipo": limpiar(eq)})
    
    for fila in range(8, 38, 4):
        eq = grilla.get((fila, "L"), "")
        if eq:
            bracket_izq["8avos"].append({"equipo": limpiar(eq)})
    
    for fila in range(10, 38, 8):
        eq = grilla.get((fila, "N"), "")
        if eq:
            bracket_izq["cuartos"].append({"equipo": limpiar(eq)})
    
    for fila in (14, 30):
        eq = grilla.get((fila, "P"), "")
        if eq:
            bracket_izq["semis"].append({"equipo": limpiar(eq)})
    
    # ── Bracket RIGHT: AE(slots)/AD(16avos)/AB(8avos)/Z(cuartos)/X(semis) ──
    bracket_der = {"16avos": [], "8avos": [], "cuartos": [], "semis": []}
    for fila in range(7, 40, 2):
        slot = grilla.get((fila, "AE"), "")
        eq = grilla.get((fila, "AD"), "")
        if slot and es_slot_code(slot):
            bracket_der["16avos"].append({"slot": slot, "equipo": limpiar(eq)})
    
    for fila in range(8, 38, 4):
        eq = grilla.get((fila, "AB"), "")
        if eq:
            bracket_der["8avos"].append({"equipo": limpiar(eq)})
    
    for fila in range(10, 38, 8):
        eq = grilla.get((fila, "Z"), "")
        if eq:
            bracket_der["cuartos"].append({"equipo": limpiar(eq)})
    
    for fila in (14, 30):
        eq = grilla.get((fila, "X"), "")
        if eq:
            bracket_der["semis"].append({"equipo": limpiar(eq)})
    
    # Consolidar
    ronda_16 = bracket_izq["16avos"] + bracket_der["16avos"]
    ronda_8 = bracket_izq["8avos"] + bracket_der["8avos"]
    ronda_4 = bracket_izq["cuartos"] + bracket_der["cuartos"]
    ronda_2 = bracket_izq["semis"] + bracket_der["semis"]
    
    # ── Finales ──
    # S16 = CAMPEÓN label, S17 = campeón team
    # Also look for "CAMPEÓN" in R, T, etc.
    campeon = limpiar(grilla.get((17, "S"), ""))
    if not campeon:
        for (fila, col), val in grilla.items():
            if "CAMPEÓN" in val.upper():
                for offset in [(1,0), (0,-1), (0,1), (-1,0)]:
                    c = chr(ord(col) + offset[1])
                    eq = limpiar(grilla.get((fila + offset[0], c), ""))
                    if eq:
                        campeon = eq
                        break
                if campeon:
                    break
    
    # Buscar otros puestos (2°, 3°, 4°) en columnas R/S/T/U
    puestos = {}
    for fila in range(15, 30):
        for col in ("R", "S", "T", "U"):
            val = grilla.get((fila, col), "")
            if val and not val.upper().startswith(("CAMPEÓN", "FINAL", "SEMI", "CUARTOS")):
                eq = limpiar(val)
                # Buscar etiqueta NUMÉRICA cercana (hacia arriba o misma celda)
                for offset_col in [-2, -1, 0]:
                    col_num = chr(ord(col) + offset_col) if ord(col) + offset_col >= ord('A') else 'A'
                    num_val = grilla.get((fila, col_num), "")
                    try:
                        pos = int(float(num_val))
                        if pos in (1,2,3,4):
                            nombre = {1: "campeon", 2: "segundo", 3: "tercero", 4: "cuarto"}.get(pos)
                            if nombre and nombre not in puestos:
                                puestos[nombre] = eq
                    except:
                        pass
    
    # Si ya tenemos campeon por S17, priorizarlo
    if campeon and "campeon" not in puestos:
        puestos["campeon"] = campeon
    
    finales = {
        "campeon": puestos.get("campeon"),
        "segundo": puestos.get("segundo"),
        "tercero": puestos.get("tercero"),
        "cuarto": puestos.get("cuarto"),
    }
    
    # ── Validar ──
    errores = []
    if len(grupos) != 12:
        errores.append(f"Se encontraron {len(grupos)} grupos, se esperaban 12")
    
    eq16 = sum(1 for e in ronda_16 if e["equipo"])
    if eq16 != 32:
        errores.append(f"16avos: {eq16} equipos (se esperaban 32)")
    if len(ronda_8) != 16:
        errores.append(f"8avos: {len(ronda_8)} equipos (se esperaban 16)")
    if len(ronda_4) != 8:
        errores.append(f"Cuartos: {len(ronda_4)} equipos (se esperaban 8)")
    if len(ronda_2) != 4:
        errores.append(f"Semifinales: {len(ronda_2)} equipos (se esperaban 4)")
    
    puestos_falt = [k for k, v in finales.items() if not v]
    if puestos_falt:
        errores.append(f"Finales incompletas: {', '.join(puestos_falt)}")
    
    return {
        "archivo_original": os.path.basename(filepath),
        "participante": limpiar(participante),
        "grupos": grupos,
        "ronda_16avos": ronda_16,
        "ronda_8avos": ronda_8,
        "ronda_cuartos": ronda_4,
        "ronda_semifinales": ronda_2,
        "finales": finales,
        "errores": errores,
    }

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 importar_excel_v3.py <archivo.xlsx> [...]")
        sys.exit(1)
    salida = Path("data/pollas")
    salida.mkdir(parents=True, exist_ok=True)
    for fp in sys.argv[1:]:
        print(f"\n{'='*60}\nParseando (v3): {fp}")
        try:
            d = parsear_excel_v3(fp)
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback; traceback.print_exc()
            continue
        print(f"  Participante: {d['participante']}")
        print(f"  Grupos: {len(d['grupos'])}/12")
        print(f"  16avos: {sum(1 for e in d['ronda_16avos'] if e['equipo'])}/32")
        print(f"  8avos: {len(d['ronda_8avos'])}/16")
        print(f"  Cuartos: {len(d['ronda_cuartos'])}/8")
        print(f"  Semis: {len(d['ronda_semifinales'])}/4")
        print(f"  🏆 {d['finales']['campeon']} | 🥈 {d['finales']['segundo']} | 🥉 {d['finales']['tercero']} | 4° {d['finales']['cuarto']}")
        if d["errores"]:
            for e in d["errores"]:
                print(f"  ⚠️  {e}")
        else:
            print("  ✅ Sin errores")
        nombre = Path(fp).stem.lower().replace(" ", "_")
        with open(salida / f"{nombre}.json", "w", encoding="utf-8") as f:
            json.dump(d, f, indent=2, ensure_ascii=False)
        print(f"  💾 {salida / f'{nombre}.json'}")

if __name__ == "__main__":
    main()
