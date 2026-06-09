#!/usr/bin/env python3
"""
Parser v2 para pollas en formato horizontal (Fer, Seb, etc.)

Estructura:
- Grupos: 12 grupos en columnas B-Y, filas 5-8
- Bracket IZQUIERDO: slots en columna C, equipos en columna D (filas pares 12-42)
- Bracket DERECHO: slots en columna W, equipos en columna V (filas pares 12-42)
- 8avos IZQ: columna F | 8avos DER: columna T
- Cuartos IZQ: columna H | Cuartos DER: columna R  
- Semifinales: columnas J, P (filas 19, 35)
- Finales: posiciones 1-4 en K38-L41

Uso: python3 importar_excel_v2.py <archivo.xlsx> [...]
"""

import json, os, re, sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl no instalado. pip3 install openpyxl")
    sys.exit(1)

def limpiar(texto):
    return " ".join(str(texto).split()).strip() if texto else ""

def es_slot(texto):
    """Detecta código de llave: 1E, 2A, 3 ABCDF, 3 CEFHI, etc."""
    t = limpiar(texto).upper()
    return bool(re.match(r'^[123]\s*[A-L]+$', t))

def parsear_excel_v2(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Construir grilla {(fila, col_letra): valor}
    grilla = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 60):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    grilla[(cell.row, cell.column_letter)] = val
    
    # ── Participante (D2) ──
    participante = grilla.get((2, "D"), "Sin nombre")
    
    # ── Grupos (horizontal B-Y, rows 5-8) ──
    grupos = {}
    # Row 4 headers: "GRUPO A" at B, "GRUPO B" at D, etc. 
    # Teams: rows 5-8, even columns (B,D,F,...), positions: odd columns (C,E,G,...)
    col_team_start = ord('B') - ord('A')
    col_team_end = ord('Y') - ord('A')
    
    for col_idx in range(col_team_start, col_team_end + 1, 2):
        col_e = chr(ord('A') + col_idx)
        col_p = chr(ord('A') + col_idx + 1)
        
        header = grilla.get((4, col_e), "")
        if not header.upper().startswith("GRUPO"):
            continue
        letra = header.replace("GRUPO", "").strip()
        if not letra or len(letra) > 2:
            continue
        
        equipos = {}
        for fila in range(5, 9):
            eq = limpiar(grilla.get((fila, col_e), ""))
            pos_str = grilla.get((fila, col_p), "")
            if not eq:
                continue
            try:
                pos = int(float(pos_str))
            except (ValueError, TypeError):
                pos = None
            equipos[eq] = pos
        
        if equipos:
            grupos[letra] = equipos
    
    # ── 16avos: LEFT (C+D) + RIGHT (W+V), rows 12-42 (even rows) ──
    def extraer_16avos(col_slot, col_team):
        """Extrae slots y equipos de columnas específicas."""
        slots = []
        for fila in range(12, 44, 2):  # 12,14,16,...,42
            slot = grilla.get((fila, col_slot), "")
            if es_slot(slot):
                eq = limpiar(grilla.get((fila, col_team), ""))
                slots.append({"slot": slot, "equipo": eq})
        return slots
    
    ronda_16 = extraer_16avos("C", "D") + extraer_16avos("W", "V")
    
    # ── 8avos: LEFT (F) + RIGHT (T), equipos en filas específicas ──
    def extraer_ronda(col, filas):
        equipos = []
        for f in filas:
            eq = limpiar(grilla.get((f, col), ""))
            if eq and not es_slot(eq) and not re.match(r'^M\d{2,3}', eq):
                equipos.append({"equipo": eq})
        return equipos
    
    ronda_8 = extraer_ronda("F", [13,17,21,25,29,33,37,41]) + \
              extraer_ronda("T", [13,17,21,25,29,33,37,41])
    
    # ── Cuartos: LEFT (H) + RIGHT (R) ──
    ronda_4 = extraer_ronda("H", [15,23,31,39]) + \
              extraer_ronda("R", [15,23,31,39])
    
    # ── Semifinales: LEFT (J) + RIGHT (P) ──
    ronda_2 = extraer_ronda("J", [19,35]) + \
              extraer_ronda("P", [19,35])
    
    # ── Finales: posiciones 1-4 en K38-L41 ──
    posiciones = {}
    for fila in range(38, 42):
        pos_str = grilla.get((fila, "K"), "")
        equipo = limpiar(grilla.get((fila, "L"), ""))
        try:
            pos = int(float(pos_str))
        except (ValueError, TypeError):
            continue
        nombre = {1: "campeon", 2: "segundo", 3: "tercero", 4: "cuarto"}.get(pos)
        if nombre:
            posiciones[nombre] = equipo
    
    # Si no encontramos posiciones en K-L, buscar en otras celdas
    if not posiciones:
        # Buscar labels FINAL, 3° LUGAR, etc.
        for (fila, col), val in grilla.items():
            if "CAMPEÓN" in val.upper() or "FINAL" in val.upper():
                # Buscar equipos cercanos
                pass
    
    finales = {
        "campeon": posiciones.get("campeon"),
        "segundo": posiciones.get("segundo"),
        "tercero": posiciones.get("tercero"),
        "cuarto": posiciones.get("cuarto"),
    }
    
    # ── Validar ──
    errores = []
    if len(grupos) != 12:
        errores.append(f"Se encontraron {len(grupos)} grupos, se esperaban 12")
    
    equipos_16 = [e["equipo"] for e in ronda_16 if e["equipo"]]
    if len(equipos_16) != 32:
        errores.append(f"16avos: {len(equipos_16)} equipos (se esperaban 32)")
    
    if len(ronda_8) != 16:
        errores.append(f"8avos: {len(ronda_8)} equipos (se esperaban 16)")
    
    if len(ronda_4) != 8:
        errores.append(f"Cuartos: {len(ronda_4)} equipos (se esperaban 8)")
    
    if len(ronda_2) != 4:
        errores.append(f"Semifinales: {len(ronda_2)} equipos (se esperaban 4)")
    
    puestos_faltantes = [k for k, v in finales.items() if v is None]
    if puestos_faltantes:
        errores.append(f"Finales incompletas: faltan {', '.join(puestos_faltantes)}")
    
    return {
        "archivo_original": os.path.basename(filepath),
        "participante": limpiar(participante),
        "grupos": grupos,
        "ronda_16avos": ronda_16,
        "ronda_8avos": ronda_8,
        "ronda_cuartos": ronda_4,
        "ronda_semifinales": ronda_2,
        "finales": finales,
        "errores": errores,
    }

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 importar_excel_v2.py <archivo.xlsx> [...]")
        sys.exit(1)
    
    salida_dir = Path("data/pollas")
    salida_dir.mkdir(parents=True, exist_ok=True)
    
    for fp in sys.argv[1:]:
        if not os.path.exists(fp):
            print(f"⚠️  No existe: {fp}")
            continue
        print(f"\n{'='*60}")
        print(f"Parseando (v2): {fp}")
        
        try:
            datos = parsear_excel_v2(fp)
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback; traceback.print_exc()
            continue
        
        print(f"  Participante: {datos['participante']}")
        print(f"  Grupos: {len(datos['grupos'])}/12")
        eq16 = sum(1 for e in datos['ronda_16avos'] if e['equipo'])
        print(f"  16avos: {eq16}/32")
        print(f"  8avos: {len(datos['ronda_8avos'])}/16")
        print(f"  Cuartos: {len(datos['ronda_cuartos'])}/8")
        print(f"  Semis: {len(datos['ronda_semifinales'])}/4")
        print(f"  🏆 Campeón: {datos['finales']['campeon']}")
        print(f"  🥈 Segundo: {datos['finales']['segundo']}")
        print(f"  🥉 Tercero: {datos['finales']['tercero']}")
        print(f"  4to: {datos['finales']['cuarto']}")
        
        if datos["errores"]:
            print(f"\n  ⚠️  Errores:")
            for e in datos["errores"]:
                print(f"     - {e}")
        else:
            print(f"  ✅ Sin errores")
        
        nombre = Path(fp).stem.lower().replace(" ", "_")
        with open(salida_dir / f"{nombre}.json", "w", encoding="utf-8") as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        print(f"  💾 {salida_dir / f'{nombre}.json'}")

if __name__ == "__main__":
    main()
