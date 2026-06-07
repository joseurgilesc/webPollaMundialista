#!/usr/bin/env python3
"""
Parser de pollas mundialistas desde archivos Excel (.xlsx).

Extrae:
- Fase de grupos (posiciones 1-4 por equipo)
- Bracket de eliminación (dos mitades: izquierda H-O, derecha AD-W)
- Finales (campeón, segundo, tercero, cuarto)

Uso: python3 importar_excel.py [archivo.xlsx ...]
     python3 importar_excel.py data/pollas/*.xlsx
"""

import json
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl no está instalado. Ejecutá: pip3 install openpyxl")
    sys.exit(1)


# ── Mapeo de columnas ──────────────────────────────────────────────

# Grupos: pares de 2 por fila de encabezado
#   A/B → columnas A-B, C/D → A-B también pero en otra fila, alternan A/E
#   Grupo izquierdo: col A (nombre), col B (posición)
#   Grupo derecho:   col D (nombre), col E (posición)

# Bracket izquierdo (cols H-O)
BRACKET_IZQ = {
    "slots":    "H",   # códigos: 1E, 3 ABCDF, 2A, ...
    "16avos":   "I",   # equipos en 16avos
    "8avos":    "K",   # ganadores de 16avos
    "cuartos":  "M",   # ganadores de 8avos
    "semis":    "O",   # ganadores de cuartos
}

# Bracket derecho (cols AD-W)
BRACKET_DER = {
    "slots":    "AD",  # códigos
    "16avos":   "AC",  # equipos en 16avos
    "8avos":    "AA",  # ganadores de 16avos
    "cuartos":  "Y",   # ganadores de 8avos
    "semis":    "W",   # ganadores de cuartos
}

# Finales (cols Q-U)
FINALES = {
    "campeon":              "R17",
    "finalista_izq":        "Q20",
    "finalista_der":        "T20",
    "tercero":              "R29",
    "semi_perdedor_izq":    "Q26",
    "semi_perdedor_der":    "T26",
}

# ── Utilidades ─────────────────────────────────────────────────────

def col_letra_a_numero(col: str) -> int:
    """Convierte letra de columna Excel a índice 0-based."""
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n - 1


def celda_a_fila_col(ref: str) -> tuple[int, str]:
    """Separa 'R17' → (17, 'R')."""
    col = "".join(c for c in ref if c.isalpha())
    fila = int("".join(c for c in ref if c.isdigit()))
    return fila, col


def limpiar_equipo(texto: str) -> str:
    """Normaliza nombres de equipos: quita espacios extra, banderas."""
    if not texto:
        return ""
    # Quitar saltos de línea y espacios múltiples
    texto = " ".join(texto.split())
    return texto.strip()


def es_slot_code(texto: str) -> bool:
    """Detecta si un texto es un código de llave (1E, 2A, 3 ABCDF, etc.)."""
    if not texto:
        return False
    texto = texto.strip().upper()
    # Patrones: 1E, 12L, 2A, 3 ABCDF, 3 CDFGH, etc.
    import re
    return bool(re.match(r'^[123]\s*[A-L]+$', texto))


# ── Parser principal ───────────────────────────────────────────────

def parsear_excel(filepath: str) -> dict:
    """
    Parsea un archivo Excel de polla mundialista.
    Retorna un dict con la estructura completa.
    """
    wb = load_workbook(filepath, data_only=True)
    
    if len(wb.sheetnames) < 1:
        raise ValueError(f"El archivo {filepath} no tiene hojas.")
    
    ws = wb[wb.sheetnames[0]]  # Hoja 1 = predicciones
    
    # Construir grilla: {(fila, col_letra): valor}
    grilla = {}
    max_fila = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 60):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    col_letra = cell.column_letter
                    grilla[(cell.row, col_letra)] = val
                    max_fila = max(max_fila, cell.row)
    
    nombre_participante = _extraer_participante(grilla)
    grupos = _extraer_grupos(grilla)
    bracket_izq = _extraer_bracket_lado(grilla, BRACKET_IZQ, "izquierdo")
    bracket_der = _extraer_bracket_lado(grilla, BRACKET_DER, "derecho")
    finales = _extraer_finales(grilla)
    
    # Consolidar bracket
    ronda_16avos = bracket_izq["16avos"] + bracket_der["16avos"]
    ronda_8avos = bracket_izq["8avos"] + bracket_der["8avos"]
    ronda_cuartos = bracket_izq["cuartos"] + bracket_der["cuartos"]
    ronda_semis = bracket_izq["semis"] + bracket_der["semis"]
    
    # Errores detectados
    errores = _validar(grupos, ronda_16avos, ronda_8avos, ronda_cuartos, ronda_semis, finales)
    
    return {
        "archivo_original": os.path.basename(filepath),
        "participante": nombre_participante,
        "grupos": grupos,
        "ronda_16avos": ronda_16avos,
        "ronda_8avos": ronda_8avos,
        "ronda_cuartos": ronda_cuartos,
        "ronda_semifinales": ronda_semis,
        "finales": finales,
        "errores": errores,
    }


def _extraer_participante(grilla: dict) -> str:
    """Extrae el nombre del participante de la celda F2 (o similar)."""
    # Buscar en F2 primero (formato SCC2, SCC3, SCC4)
    if (2, "F") in grilla:
        nombre = grilla[(2, "F")]
        if nombre and "NOMBRES" not in nombre.upper():
            return limpiar_equipo(nombre)
    
    # Buscar patrón "NOMBRES Y APELLIDOS:" y tomar la celda a la derecha
    for (fila, col), val in grilla.items():
        if "NOMBRES" in val.upper() and "APELLIDOS" in val.upper():
            # Buscar en F de la misma fila
            if (fila, "F") in grilla:
                return limpiar_equipo(grilla[(fila, "F")])
    
    return "Sin nombre"


def _extraer_grupos(grilla: dict) -> dict:
    """
    Extrae los 12 grupos (A-L) con sus 4 equipos y posiciones.
    
    Formato en el Excel:
      A5=GRUPO A  B5=#    D5=GRUPO B  E5=#
      A6=equipo   B6=pos  D6=equipo   E6=pos
      ...4 filas de equipos...
      A11=GRUPO C  B11=#   D11=GRUPO D  E11=#
    """
    grupos = {}
    
    # Buscar filas que contengan "GRUPO X" en col A o D
    filas_grupo = []
    for (fila, col), val in grilla.items():
        if col in ("A", "D") and val.upper().startswith("GRUPO"):
            filas_grupo.append((fila, col, val))
    
    for fila_encabezado, col_encabezado, texto in filas_grupo:
        # Extraer letra del grupo
        letra = texto.replace("GRUPO", "").strip()
        if not letra or len(letra) > 2:
            continue
        
        # Determinar columnas según lado
        if col_encabezado == "A":
            col_equipo = "A"
            col_pos = "B"
        else:
            col_equipo = "D"
            col_pos = "E"
        
        equipos = {}
        for offset in range(1, 5):
            f = fila_encabezado + offset
            equipo = grilla.get((f, col_equipo), "")
            pos_str = grilla.get((f, col_pos), "")
            
            if not equipo:
                continue
            
            equipo = limpiar_equipo(equipo)
            try:
                pos = int(float(pos_str))
            except (ValueError, TypeError):
                pos = None
            
            equipos[equipo] = pos
        
        if equipos:
            grupos[letra] = equipos
    
    return grupos


def _extraer_bracket_lado(grilla: dict, columnas: dict, lado: str) -> dict:
    """
    Extrae un lado del bracket (izquierdo o derecho).
    
    El bracket está organizado por filas:
    - Las filas impares desde 7 (7,9,11,...,37) tienen slots + equipos 16avos
    - Los ganadores avanzan a columnas de rondas posteriores en filas pares
    """
    col_slots = columnas["slots"]
    col_16 = columnas["16avos"]
    col_8 = columnas["8avos"]
    col_cuartos = columnas["cuartos"]
    col_semis = columnas["semis"]
    
    resultado = {
        "16avos": [],
        "8avos": [],
        "cuartos": [],
        "semis": [],
    }
    
    # ── 16avos: slots + equipos ──
    for fila in range(7, 40, 2):  # 7, 9, 11, ..., 37
        slot = grilla.get((fila, col_slots), "")
        equipo = grilla.get((fila, col_16), "")
        
        if slot or equipo:
            resultado["16avos"].append({
                "slot": limpiar_equipo(slot) if slot else None,
                "equipo": limpiar_equipo(equipo) if equipo else None,
            })
    
    # ── 8avos ──
    for fila in range(8, 38, 4):  # 8, 12, 16, 20, 24, 28, 32, 36
        equipo = grilla.get((fila, col_8), "")
        if equipo:
            resultado["8avos"].append({
                "equipo": limpiar_equipo(equipo),
            })
    
    # ── Cuartos ──
    for fila in range(10, 38, 8):  # 10, 18, 26, 34
        equipo = grilla.get((fila, col_cuartos), "")
        if equipo:
            resultado["cuartos"].append({
                "equipo": limpiar_equipo(equipo),
            })
    
    # ── Semifinales ──
    for fila in (14, 30):
        equipo = grilla.get((fila, col_semis), "")
        if equipo:
            resultado["semis"].append({
                "equipo": limpiar_equipo(equipo),
            })
    
    return resultado


def _extraer_finales(grilla: dict) -> dict:
    """
    Extrae los 4 puestos finales.
    
    Estructura:
    - R17 = CAMPEÓN
    - Q20 = finalista izquierdo, T20 = finalista derecho
      → segundo = el de Q20/T20 que NO sea R17
    - R29 = TERCERO
    - Q26 = semi-perdedor izq, T26 = semi-perdedor der
      → cuarto = el de Q26/T26 que NO sea R29
    """
    def celda(ref: str) -> str:
        f, c = celda_a_fila_col(ref)
        return limpiar_equipo(grilla.get((f, c), ""))
    
    campeon = celda("R17")
    fin_izq = celda("Q20")
    fin_der = celda("T20")
    tercero = celda("R29")
    semi_perd_izq = celda("Q26")
    semi_perd_der = celda("T26")
    
    # Deducir segundo
    if campeon:
        if fin_izq and fin_izq != campeon:
            segundo = fin_izq
        elif fin_der and fin_der != campeon:
            segundo = fin_der
        else:
            segundo = None
    else:
        segundo = None
    
    # Deducir cuarto
    if tercero:
        if semi_perd_izq and semi_perd_izq != tercero:
            cuarto = semi_perd_izq
        elif semi_perd_der and semi_perd_der != tercero:
            cuarto = semi_perd_der
        else:
            cuarto = None
    else:
        cuarto = None
    
    return {
        "campeon": campeon or None,
        "segundo": segundo,
        "tercero": tercero or None,
        "cuarto": cuarto,
    }


# ── Validación ─────────────────────────────────────────────────────

def _validar(grupos: dict, r16: list, r8: list, r4: list, r2: list, finales: dict) -> list:
    """Detecta problemas en los datos parseados."""
    errores = []
    
    # Grupos: deben ser 12
    if len(grupos) != 12:
        errores.append(f"Se encontraron {len(grupos)} grupos, se esperaban 12.")
    
    # Cada grupo debe tener 4 equipos con posiciones 1-4
    for letra, equipos in grupos.items():
        if len(equipos) != 4:
            errores.append(f"Grupo {letra}: {len(equipos)} equipos (se esperaban 4).")
        for equipo, pos in equipos.items():
            if pos is None:
                errores.append(f"Grupo {letra}: '{equipo}' no tiene posición numérica.")
    
    # 16avos: deben ser 32 entradas (16 por lado)
    entradas_con_datos = [e for e in r16 if e["equipo"]]
    if len(entradas_con_datos) != 32:
        errores.append(f"16avos: {len(entradas_con_datos)} equipos (se esperaban 32).")
    
    # 8avos: 16 equipos
    if len(r8) != 16:
        errores.append(f"8avos: {len(r8)} equipos (se esperaban 16).")
    
    # Cuartos: 8 equipos
    if len(r4) != 8:
        errores.append(f"Cuartos: {len(r4)} equipos (se esperaban 8).")
    
    # Semis: 4 equipos
    if len(r2) != 4:
        errores.append(f"Semifinales: {len(r2)} equipos (se esperaban 4).")
    
    # Finales: 4 puestos
    puestos_faltantes = [k for k, v in finales.items() if v is None]
    if puestos_faltantes:
        errores.append(f"Finales incompletas: faltan {', '.join(puestos_faltantes)}.")
    
    return errores


# ── CLI ────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 importar_excel.py <archivo.xlsx> [...]")
        print("     python3 importar_excel.py data/pollas/*.xlsx")
        sys.exit(1)
    
    salida_dir = Path("data/pollas")
    salida_dir.mkdir(parents=True, exist_ok=True)
    
    for filepath in sys.argv[1:]:
        if not os.path.exists(filepath):
            print(f"⚠️  No existe: {filepath}")
            continue
        
        if not filepath.lower().endswith(".xlsx"):
            print(f"⚠️  No es Excel, salteando: {filepath}")
            continue
        
        print(f"\n{'='*60}")
        print(f"Parseando: {filepath}")
        
        try:
            datos = parsear_excel(filepath)
        except Exception as e:
            print(f"❌ Error al parsear: {e}")
            import traceback
            traceback.print_exc()
            continue
        
        # Mostrar resumen
        print(f"  Participante: {datos['participante']}")
        print(f"  Grupos: {len(datos['grupos'])}")
        print(f"  16avos: {len(datos['ronda_16avos'])} equipos")
        print(f"  8avos: {len(datos['ronda_8avos'])} equipos")
        print(f"  Cuartos: {len(datos['ronda_cuartos'])} equipos")
        print(f"  Semis: {len(datos['ronda_semifinales'])} equipos")
        print(f"  🏆 Campeón: {datos['finales']['campeon']}")
        print(f"  🥈 Segundo: {datos['finales']['segundo']}")
        print(f"  🥉 Tercero: {datos['finales']['tercero']}")
        print(f"  4to: {datos['finales']['cuarto']}")
        
        if datos["errores"]:
            print(f"\n  ⚠️  Errores detectados:")
            for e in datos["errores"]:
                print(f"     - {e}")
        else:
            print(f"  ✅ Sin errores")
        
        # Guardar JSON
        nombre_base = Path(filepath).stem.lower().replace(" ", "_")
        salida = salida_dir / f"{nombre_base}.json"
        with open(salida, "w", encoding="utf-8") as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        print(f"  💾 Guardado en: {salida}")


if __name__ == "__main__":
    main()
